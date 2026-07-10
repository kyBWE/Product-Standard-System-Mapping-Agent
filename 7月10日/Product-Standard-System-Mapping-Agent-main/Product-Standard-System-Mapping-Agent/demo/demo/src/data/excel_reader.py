from __future__ import annotations
import ast
import json
import logging
import os
from datetime import datetime

from openpyxl import load_workbook, Workbook

from src.models.category_node import CategoryNode
from src.models.evolve_models import SynonymUpdateAction


logger = logging.getLogger("ExcelDataReader")


class ExcelDataReader:
    def load_standard_system(self, file_path: str) -> tuple[list[CategoryNode], int]:
        if not os.path.exists(file_path):
            logger.error(f"标准体系文件不存在: {file_path}")
            raise FileNotFoundError(f"标准体系文件不存在: {file_path}")

        wb = load_workbook(file_path, read_only=True)
        ws = wb.active
        nodes: list[CategoryNode] = []
        skipped = 0

        header_row = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        col_map = {str(h).strip(): i for i, h in enumerate(header_row) if h is not None}

        required = {"category_id", "category_name"}
        missing_cols = required - set(col_map.keys())
        if missing_cols:
            wb.close()
            raise ValueError(f"标准体系文件缺少必需列: {missing_cols}")

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                category_id = str(row[col_map["category_id"]]).strip() if row[col_map.get("category_id", -1)] is not None else ""
                category_name = str(row[col_map["category_name"]]).strip() if row[col_map.get("category_name", -1)] is not None else ""

                if not category_id or not category_name:
                    skipped += 1
                    logger.warning(f"第{row_idx}行缺少必需字段, 已跳过")
                    continue

                category_pids_raw = row[col_map.get("category_pids", -1)]
                category_pids = self._parse_pids_field(category_pids_raw)
                category_group_name = str(row[col_map.get("category_group_name", -1)] or "").strip()
                syn_list = self._parse_list_field(row[col_map.get("syn_list", -1)])

                nodes.append(CategoryNode(
                    category_id=category_id,
                    category_name=category_name,
                    category_pids=category_pids,
                    category_group_name=category_group_name,
                    syn_list=syn_list,
                ))
            except Exception as e:
                skipped += 1
                logger.warning(f"第{row_idx}行解析异常: {e}, 已跳过")

        wb.close()
        logger.info(f"标准体系加载完成: 共{len(nodes)}条, 跳过{skipped}条")
        return nodes, skipped

    def load_company_products(self, file_path: str) -> list[str]:
        if not os.path.exists(file_path):
            logger.error(f"企业产品文件不存在: {file_path}")
            raise FileNotFoundError(f"企业产品文件不存在: {file_path}")

        wb = load_workbook(file_path, read_only=True)
        ws = wb.active
        products: list[str] = []

        header_row = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        col_map = {str(h).strip(): i for i, h in enumerate(header_row) if h is not None}

        if "product_name" not in col_map:
            wb.close()
            raise ValueError("企业产品文件缺少 product_name 列")

        pn_idx = col_map["product_name"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[pn_idx] is not None:
                name = str(row[pn_idx]).strip()
                if name:
                    products.append(name)

        wb.close()
        logger.info(f"企业产品数据加载完成: 共{len(products)}条")
        return products

    def count_company_products(self, file_path: str) -> int:
        if not os.path.exists(file_path):
            logger.error(f"企业产品文件不存在: {file_path}")
            raise FileNotFoundError(f"企业产品文件不存在: {file_path}")

        wb = load_workbook(file_path, read_only=True)
        ws = wb.active

        header_row = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        col_map = {str(h).strip(): i for i, h in enumerate(header_row) if h is not None}

        if "product_name" not in col_map:
            wb.close()
            raise ValueError("企业产品文件缺少 product_name 列")

        pn_idx = col_map["product_name"]
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[pn_idx] is not None and str(row[pn_idx]).strip():
                count += 1

        wb.close()
        logger.info(f"企业产品数据计数完成: 共{count}条")
        return count

    def write_back_synonyms(self, file_path: str, updates: list[SynonymUpdateAction]) -> int:
        if not os.path.exists(file_path):
            logger.error(f"标准体系文件不存在, 无法回写: {file_path}")
            return 0

        wb = load_workbook(file_path)
        ws = wb.active

        header_row = [cell.value for cell in ws[1]]
        col_map = {str(h).strip(): i + 1 for i, h in enumerate(header_row) if h is not None}

        id_col = col_map.get("category_id")
        syn_col = col_map.get("syn_list")
        if id_col is None or syn_col is None:
            wb.close()
            logger.error("文件缺少 category_id 或 syn_list 列, 无法回写")
            return 0

        id_row_map: dict[str, int] = {}
        for row_idx in range(2, ws.max_row + 1):
            cell_val = ws.cell(row=row_idx, column=id_col).value
            if cell_val is not None:
                id_row_map[str(cell_val).strip()] = row_idx

        success_count = 0
        for update in updates:
            row_idx = id_row_map.get(update.category_id)
            if row_idx is None:
                logger.warning(f"回写同义词: 未找到category_id={update.category_id}")
                continue

            current_syn_cell = ws.cell(row=row_idx, column=syn_col)
            current_syn = self._parse_list_field(current_syn_cell.value)

            if update.product_name in current_syn:
                logger.info(f"同义词已存在, 跳过: category_id={update.category_id}, synonym={update.product_name}")
                continue

            current_syn.append(update.product_name)
            ws.cell(row=row_idx, column=syn_col).value = json.dumps(current_syn, ensure_ascii=False)
            success_count += 1

        try:
            wb.save(file_path)
            logger.info(f"同义词回写完成: 成功{success_count}条")
        except Exception as e:
            logger.error(f"同义词回写保存失败: {e}")
        finally:
            wb.close()

        return success_count

    @staticmethod
    def _parse_list_field(value: object) -> list[str]:
        if value is None:
            return []
        if isinstance(value, list):
            return [str(v).strip() for v in value if v is not None and str(v).strip()]
        text = str(value).strip()
        if not text:
            return []
        try:
            parsed = ast.literal_eval(text)
            if isinstance(parsed, list):
                return [str(v).strip() for v in parsed if v is not None and str(v).strip()]
            return [text]
        except (ValueError, SyntaxError):
            if "," in text:
                return [v.strip() for v in text.split(",") if v.strip()]
            if "、" in text:
                return [v.strip() for v in text.split("、") if v.strip()]
            return [text] if text else []

    @staticmethod
    def _parse_pids_field(value: object) -> list[str]:
        """解析category_pids字段，格式如 '[-1],[2],[3],[4]'，过滤虚拟根节点-1"""
        if value is None:
            return []
        text = str(value).strip()
        if not text:
            return []
        import re
        brackets = re.findall(r'\[([^\]]*)\]', text)
        result = [b.strip() for b in brackets if b.strip() and b.strip() != "-1"]
        if not result and "," in text:
            parts = [p.strip().strip("[]") for p in text.split(",")]
            result = [p for p in parts if p and p != "-1"]
        return result
