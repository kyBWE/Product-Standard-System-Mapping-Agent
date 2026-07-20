#!/usr/bin/env python3
"""生成一份极简 Excel，演示树状有序 / 末尾乱序 / 树状插入。"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

OUT = Path(__file__).resolve().parent.parent / "output" / "树状排序与插入示例.xlsx"

HEADERS = [
    "category_id",
    "category_name",
    "category_group_id",
    "category_pids",
    "category_group_name",
    "syn_list",
    "说明",
]

ORDERED = [
    (2, "农林牧渔", None, "[-1],", None, "[]", "根"),
    (3, "农业产品", "2", "[-1],[2]", "农林牧渔", "[]", "└ 子"),
    (4, "谷物", "2,3", "[-1],[2],[3]", "农林牧渔,农业产品", "[]", "   └ 子"),
    (5, "稻谷", "2,3,4", "[-1],[2],[3],[4]", "农林牧渔,农业产品,谷物", "[]", "      └ 子"),
    (6, "早籼稻", "2,3,4,5", "[-1],[2],[3],[4],[5]", "农林牧渔,农业产品,谷物,稻谷", "[]", "         └ 叶"),
    (7, "晚籼稻", "2,3,4,5", "[-1],[2],[3],[4],[5]", "农林牧渔,农业产品,谷物,稻谷", "[]", "         └ 叶"),
    (12, "小麦", "2,3,4", "[-1],[2],[3],[4]", "农林牧渔,农业产品,谷物", "[]", "      └ 子"),
    (13, "硬质小麦", "2,3,4,12", "[-1],[2],[3],[4],[12]", "农林牧渔,农业产品,谷物,小麦", "[]", "         └ 叶"),
    (14, "软质小麦", "2,3,4,12", "[-1],[2],[3],[4],[12]", "农林牧渔,农业产品,谷物,小麦", "[]", "         └ 叶"),
    (16, "玉米", "2,3,4", "[-1],[2],[3],[4]", "农林牧渔,农业产品,谷物", '["苞米"]', "      └ 子"),
]

DISORDERED = [
    (2, "农林牧渔", None, "[-1],", None, "[]", "根"),
    (3, "农业产品", "2", "[-1],[2]", "农林牧渔", "[]", ""),
    (4, "谷物", "2,3", "[-1],[2],[3]", "农林牧渔,农业产品", "[]", ""),
    (5, "稻谷", "2,3,4", "[-1],[2],[3],[4]", "农林牧渔,农业产品,谷物", "[]", ""),
    (6, "早籼稻", "2,3,4,5", "[-1],[2],[3],[4],[5]", "农林牧渔,农业产品,谷物,稻谷", "[]", ""),
    (7, "晚籼稻", "2,3,4,5", "[-1],[2],[3],[4],[5]", "农林牧渔,农业产品,谷物,稻谷", "[]", ""),
    (12, "小麦", "2,3,4", "[-1],[2],[3],[4]", "农林牧渔,农业产品,谷物", "[]", ""),
    (13, "硬质小麦", "2,3,4,12", "[-1],[2],[3],[4],[12]", "农林牧渔,农业产品,谷物,小麦", "[]", ""),
    (14, "软质小麦", "2,3,4,12", "[-1],[2],[3],[4],[12]", "农林牧渔,农业产品,谷物,小麦", "[]", ""),
    (16, "玉米", "2,3,4", "[-1],[2],[3],[4]", "农林牧渔,农业产品,谷物", '["苞米"]', ""),
    (100, "新型小麦", "2,3,4,12", "[-1],[2],[3],[4],[12]", "农林牧渔,农业产品,谷物,小麦", "[]", "← 错在末尾，父应是小麦#12"),
]

AFTER_INSERT = [
    (2, "农林牧渔", None, "[-1],", None, "[]", "根"),
    (3, "农业产品", "2", "[-1],[2]", "农林牧渔", "[]", ""),
    (4, "谷物", "2,3", "[-1],[2],[3]", "农林牧渔,农业产品", "[]", ""),
    (5, "稻谷", "2,3,4", "[-1],[2],[3],[4]", "农林牧渔,农业产品,谷物", "[]", ""),
    (6, "早籼稻", "2,3,4,5", "[-1],[2],[3],[4],[5]", "农林牧渔,农业产品,谷物,稻谷", "[]", ""),
    (7, "晚籼稻", "2,3,4,5", "[-1],[2],[3],[4],[5]", "农林牧渔,农业产品,谷物,稻谷", "[]", ""),
    (12, "小麦", "2,3,4", "[-1],[2],[3],[4]", "农林牧渔,农业产品,谷物", "[]", "父"),
    (13, "硬质小麦", "2,3,4,12", "[-1],[2],[3],[4],[12]", "农林牧渔,农业产品,谷物,小麦", "[]", ""),
    (14, "软质小麦", "2,3,4,12", "[-1],[2],[3],[4],[12]", "农林牧渔,农业产品,谷物,小麦", "[]", ""),
    (100, "新型小麦", "2,3,4,12", "[-1],[2],[3],[4],[12]", "农林牧渔,农业产品,谷物,小麦", "[]", "← 插入：小麦子树末尾"),
    (16, "玉米", "2,3,4", "[-1],[2],[3],[4]", "农林牧渔,农业产品,谷物", '["苞米"]', ""),
]


def fill_sheet(ws, rows) -> None:
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F5D50")
    new_fill = PatternFill("solid", fgColor="FFF2CC")
    thin = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        ws.append(list(row))
        note = str(row[-1] or "")
        if "←" in note:
            for cell in ws[ws.max_row]:
                cell.fill = new_fill
        for cell in ws[ws.max_row]:
            cell.border = thin

    widths = [12, 14, 16, 28, 40, 12, 30]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = width


def main() -> None:
    wb = Workbook()

    guide = wb.active
    guide.title = "阅读说明"
    guide["A1"] = "产品标准体系 · 树状排序 / 插入 小例子"
    guide["A1"].font = Font(bold=True, size=14)
    guide["A3"] = "本文件用极简数据说明三件事："
    guide["A4"] = "1.「树状有序」：父节点下方紧挨整棵子树（先序），人眼可读。"
    guide["A5"] = "2.「末尾追加后乱序」：新节点#100 只加在表尾，路径虽对，看起来却不像挂在小麦下。"
    guide["A6"] = "3.「树状插入后」：把#100 插到小麦子树最后一个后代之后（软质小麦后、玉米前）。"
    guide["A8"] = "关键字段："
    guide["A9"] = "category_pids —— 祖先链，决定逻辑树；排序/插入脚本依赖它。"
    guide["A10"] = "category_id —— 代理主键，可 max+1，不必靠近父 id。"
    guide["A12"] = "对应脚本："
    guide["A13"] = "scripts/sort_taxonomy_excel.py  —— 乱序 → 树状有序"
    guide["A14"] = "scripts/insert_taxonomy_node.py —— 按父节点子树末尾插入"
    guide.column_dimensions["A"].width = 92

    ordered = wb.create_sheet("树状有序")
    disordered = wb.create_sheet("末尾追加后乱序")
    after = wb.create_sheet("树状插入后")
    fill_sheet(ordered, ORDERED)
    fill_sheet(disordered, DISORDERED)
    fill_sheet(after, AFTER_INSERT)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"written: {OUT}")


if __name__ == "__main__":
    main()
