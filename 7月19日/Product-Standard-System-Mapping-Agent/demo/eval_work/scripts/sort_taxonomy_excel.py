#!/usr/bin/env python3
"""
按 category_pids 将《产品标准体系.xlsx》重排为树状（先序 DFS）行序。

用法：
  python scripts/sort_taxonomy_excel.py
  python scripts/sort_taxonomy_excel.py --input 产品标准体系.xlsx --output 产品标准体系_sorted.xlsx
  python scripts/sort_taxonomy_excel.py --in-place

默认写出 *._sorted.xlsx，不覆盖原文件；加 --in-place 才原地覆盖。
"""
from __future__ import annotations

import argparse
import shutil
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

# 允许直接 python scripts/xxx.py 运行
sys.path.insert(0, str(Path(__file__).resolve().parent))
from tree_excel_utils import (  # noqa: E402
    HEADERS,
    CategoryRow,
    build_forest,
    dfs_order,
)


def load_rows(path: Path) -> list[CategoryRow]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    header = [str(h).strip() if h is not None else "" for h in next(it)]
    col = {name: i for i, name in enumerate(header)}
    missing = [h for h in ("category_id", "category_name") if h not in col]
    if missing:
        wb.close()
        raise ValueError(f"缺少列: {missing}")

    rows: list[CategoryRow] = []
    for order, row in enumerate(it):
        cid = row[col["category_id"]]
        cname = row[col["category_name"]]
        if cid is None or cname is None:
            continue
        cid_s = str(cid).strip()
        cname_s = str(cname).strip()
        if not cid_s or not cname_s:
            continue
        rows.append(
            CategoryRow(
                category_id=cid_s,
                category_name=cname_s,
                category_group_id=row[col["category_group_id"]] if "category_group_id" in col else None,
                category_pids=row[col["category_pids"]] if "category_pids" in col else None,
                category_group_name=row[col["category_group_name"]] if "category_group_name" in col else None,
                syn_list=row[col["syn_list"]] if "syn_list" in col else "[]",
                source_order=order,
            )
        )
    wb.close()
    return rows


def write_rows(path: Path, rows: list[CategoryRow]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(HEADERS)
    for r in rows:
        ws.append(list(r.as_tuple()))
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def main() -> None:
    demo_dir = Path(__file__).resolve().parent.parent
    parser = argparse.ArgumentParser(description="按树路径先序重排产品标准体系 Excel")
    parser.add_argument(
        "--input",
        type=Path,
        default=demo_dir / "产品标准体系.xlsx",
        help="输入 Excel 路径",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=None,
        help="输出路径（默认在输入文件名后加 _sorted）",
    )
    parser.add_argument(
        "--in-place",
        action="store_true",
        help="原地覆盖：先备份为 .bak_时间戳，再写回原路径",
    )
    args = parser.parse_args()

    src: Path = args.input
    if not src.exists():
        raise SystemExit(f"文件不存在: {src}")

    print(f"读取: {src}")
    rows = load_rows(src)
    print(f"节点数: {len(rows)}")

    roots = build_forest(rows)
    ordered = dfs_order(roots)
    if len(ordered) != len(rows):
        print(
            f"警告: DFS 结果数 {len(ordered)} != 原表 {len(rows)}，"
            "可能存在重复 id 或挂接异常，请检查。"
        )

    if args.in_place:
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        bak = src.with_suffix(src.suffix + f".bak_{stamp}")
        shutil.copy2(src, bak)
        out = src
        print(f"已备份: {bak}")
    else:
        out = args.output or src.with_name(src.stem + "_sorted" + src.suffix)

    write_rows(out, ordered)
    print(f"已写出树序表: {out}")
    print(f"根节点数: {len(roots)}")


if __name__ == "__main__":
    main()
