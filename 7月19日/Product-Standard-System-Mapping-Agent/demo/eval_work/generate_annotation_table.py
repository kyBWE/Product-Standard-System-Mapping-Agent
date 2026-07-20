from __future__ import annotations
import json
import os
import sys

os.chdir(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.getcwd())

from src.infrastructure.config_manager import ConfigManager
from src.infrastructure.db_manager import DBConnectionManager

try:
    import openpyxl
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

INPUT_PATH = "output/test_set_200_fixed.json"
OUTPUT_PATH = "output/人工标注表_200.xlsx"


def load_category_names(db: DBConnectionManager) -> dict[str, str]:
    rows = db.execute("SELECT category_id, category_name FROM category_vectors")
    return {row["category_id"]: row["category_name"] for row in rows}


def main():
    config = ConfigManager("config.yaml")
    db_config = config.get_db_config()
    db = DBConnectionManager(db_config)
    db.initialize()

    cat_names = load_category_names(db)
    db.close()

    with open(INPUT_PATH, "r", encoding="utf-8") as f:
        data = json.load(f)

    wb = Workbook()
    ws = wb.active
    ws.title = "人工标注"

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    headers = [
        "序号",
        "产品名称",
        "RAG结果(ID)",
        "RAG结果(分类名)",
        "RAG+Rerank结果(ID)",
        "RAG+Rerank结果(分类名)",
        "PageIndex结果(ID)",
        "PageIndex结果(分类名)",
        "PageIndex_Force结果(ID)",
        "PageIndex_Force结果(分类名)",
        "原Ground Truth(ID)",
        "原Ground Truth(分类名)",
        "原GT来源",
        "四引擎是否一致",
        "人工标注结果(ID)",
        "人工标注结果(分类名)",
        "备注",
    ]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    agree_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    disagree_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for row_idx, item in enumerate(data, 2):
        results = [
            item.get("rag_result"),
            item.get("rag_rerank_result"),
            item.get("page_index_result"),
            item.get("page_index_force_result"),
        ]
        valid_results = [r for r in results if r is not None]
        all_agree = len(set(valid_results)) <= 1 if valid_results else False

        row_data = [
            row_idx - 1,
            item["product_name"],
            item.get("rag_result") or "",
            cat_names.get(item.get("rag_result"), "") if item.get("rag_result") else "",
            item.get("rag_rerank_result") or "",
            cat_names.get(item.get("rag_rerank_result"), "") if item.get("rag_rerank_result") else "",
            item.get("page_index_result") or "",
            cat_names.get(item.get("page_index_result"), "") if item.get("page_index_result") else "",
            item.get("page_index_force_result") or "",
            cat_names.get(item.get("page_index_force_result"), "") if item.get("page_index_force_result") else "",
            item.get("ground_truth") or "",
            cat_names.get(item.get("ground_truth"), "") if item.get("ground_truth") else "",
            item.get("ground_truth_source", ""),
            "是" if all_agree else "否",
            "",
            "",
            "",
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if col_idx in (1, 14):
                cell.alignment = center_align
            else:
                cell.alignment = left_align

        agree_cell = ws.cell(row=row_idx, column=14)
        agree_cell.fill = agree_fill if all_agree else disagree_fill

    col_widths = {
        1: 6,
        2: 25,
        3: 12,
        4: 30,
        5: 16,
        6: 30,
        7: 14,
        8: 30,
        9: 20,
        10: 30,
        11: 14,
        12: 30,
        13: 22,
        14: 14,
        15: 16,
        16: 30,
        17: 20,
    }
    for col, width in col_widths.items():
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "C2"

    wb.save(OUTPUT_PATH)
    print(f"已生成: {OUTPUT_PATH}")
    print(f"总条数: {len(data)}")

    agree_count = sum(
        1 for item in data
        if len(set(r for r in [item.get("rag_result"), item.get("rag_rerank_result"),
                                item.get("page_index_result"), item.get("page_index_force_result")]
                  if r is not None)) <= 1
    )
    print(f"四引擎一致: {agree_count}/{len(data)} ({agree_count/len(data)*100:.1f}%)")
    print(f"四引擎不一致: {len(data)-agree_count}/{len(data)} ({(len(data)-agree_count)/len(data)*100:.1f}%)")


if __name__ == "__main__":
    main()