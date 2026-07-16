#!/usr/bin/env python3
"""
在《产品标准体系.xlsx》中按树状结构插入新分类节点。

位置规则：插在父节点整棵子树的最后一个后代之后（父尚无子则紧跟父行）。
字段规则：category_id = 当前最大 id+1；自动拼完整 category_pids / group_*。

用法：
  python scripts/insert_taxonomy_node.py --parent-id 12 --name "新型小麦"
  python scripts/insert_taxonomy_node.py --parent-id 12 --name "新型小麦" --syn "试验麦" --dry-run
  python scripts/insert_taxonomy_node.py --parent-id 12 --name "新型小麦" --in-place

默认写出 *_inserted.xlsx；加 --in-place 则备份后覆盖原文件。
建议：表若已因末尾追加变乱，先跑 sort_taxonomy_excel.py 再插入。
"""
from __future__ import annotations

import argparse
import json
import shutil
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent))
from tree_excel_utils import (  # noqa: E402
    HEADERS,
    CategoryRow,
    build_child_fields,
    find_subtree_end_index,
    next_category_id,
    parse_pids,
)


def load_rows(path: Path) -> list[CategoryRow]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    header = [str(h).strip() if h is not None else "" for h in next(it)]
    col = {name: i for i, name in enumerate(header)}
    missing = [h for h in ("category_id", "category_name") if h not in col]
    if missing:
        wb.close()
        raise ValueError(f"缺少列: {missing}")

    rows: list[CategoryRow] = []
    for order, row in enumerate(it):
        cid = row[col["category_id"]]
        cname = row[col["category_name"]]
        if cid is None or cname is None:
            continue
        cid_s = str(cid).strip()
        cname_s = str(cname).strip()
        if not cid_s or not cname_s:
            continue
        rows.append(
            CategoryRow(
                category_id=cid_s,
                category_name=cname_s,
                category_group_id=row[col["category_group_id"]] if "category_group_id" in col else None,
                category_pids=row[col["category_pids"]] if "category_pids" in col else None,
                category_group_name=row[col["category_group_name"]] if "category_group_name" in col else None,
                syn_list=row[col["syn_list"]] if "syn_list" in col else "[]",
                source_order=order,
            )
        )
    wb.close()
    return rows


def write_rows(path: Path, rows: list[CategoryRow]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(HEADERS)
    for r in rows:
        ws.append(list(r.as_tuple()))
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def format_syn_list(syn: str | None) -> str:
    if not syn or not syn.strip():
        return "[]"
    items = [s.strip() for s in syn.split(",") if s.strip()]
    return json.dumps(items, ensure_ascii=False)


def main() -> None:
    demo_dir = Path(__file__).resolve().parent.parent
    parser = argparse.ArgumentParser(description="按树状位置向产品标准体系 Excel 插入节点")
    parser.add_argument("--input", type=Path, default=demo_dir / "产品标准体系.xlsx")
    parser.add_argument("--output", type=Path, default=None)
    parser.add_argument("--parent-id", required=True, help="父分类 category_id")
    parser.add_argument("--name", required=True, help="新分类名称")
    parser.add_argument("--syn", default="", help="同义词，逗号分隔，可选")
    parser.add_argument("--id", default="", help="指定新 id；默认 max(id)+1")
    parser.add_argument("--in-place", action="store_true", help="备份后覆盖原文件")
    parser.add_argument("--dry-run", action="store_true", help="只打印将插入的内容，不写文件")
    args = parser.parse_args()

    src: Path = args.input
    if not src.exists():
        raise SystemExit(f"文件不存在: {src}")

    rows = load_rows(src)
    parent_id = str(args.parent_id).strip()
    parent = next((r for r in rows if r.category_id == parent_id), None)
    if parent is None:
        raise SystemExit(f"找不到父节点 category_id={parent_id}")

    new_id = str(args.id).strip() if args.id else next_category_id(rows)
    if any(r.category_id == new_id for r in rows):
        raise SystemExit(f"category_id={new_id} 已存在")

    fields = build_child_fields(parent, args.name.strip(), format_syn_list(args.syn))
    new_row = CategoryRow(
        category_id=new_id,
        category_name=fields["category_name"],
        category_group_id=fields["category_group_id"],
        category_pids=fields["category_pids"],
        category_group_name=fields["category_group_name"],
        syn_list=fields["syn_list"],
    )

    insert_at = find_subtree_end_index(rows, parent_id)
    # Excel 行号（含表头）：数据下标 0 -> 第 2 行
    excel_row = insert_at + 2

    print("—— 插入预览 ——")
    print(f"父节点: #{parent.category_id} {parent.category_name}")
    print(f"父路径: {parent.category_pids}")
    print(f"新节点: #{new_row.category_id} {new_row.category_name}")
    print(f"新 pids: {new_row.category_pids}")
    print(f"新 group_id: {new_row.category_group_id}")
    print(f"新 group_name: {new_row.category_group_name}")
    print(f"插入位置: 数据下标 {insert_at}（约 Excel 第 {excel_row} 行）")
    if insert_at > 0:
        prev = rows[insert_at - 1]
        print(f"紧跟上行: #{prev.category_id} {prev.category_name}")
    if insert_at < len(rows):
        nxt = rows[insert_at]
        print(f"紧接下行: #{nxt.category_id} {nxt.category_name}")

    # 父路径是否看起来完整（仅提示）
    if not parse_pids(parent.category_pids) and parent.category_pids not in ("[-1],", "[-1]", None):
        print("注意: 父节点 pids 解析为空，请确认原表该行列是否正常。")

    if args.dry_run:
        print("dry-run：未写文件")
        return

    rows.insert(insert_at, new_row)

    if args.in_place:
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        bak = src.with_suffix(src.suffix + f".bak_{stamp}")
        shutil.copy2(src, bak)
        out = src
        print(f"已备份: {bak}")
    else:
        out = args.output or src.with_name(src.stem + "_inserted" + src.suffix)

    write_rows(out, rows)
    print(f"已写出: {out}")


if __name__ == "__main__":
    main()
