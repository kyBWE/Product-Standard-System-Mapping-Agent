#!/usr/bin/env python3
"""生成「乱序 → 树状排序 → 有序」对比示例 Excel（真的跑一遍排序脚本逻辑）。"""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

sys.path.insert(0, str(Path(__file__).resolve().parent))
from tree_excel_utils import CategoryRow, HEADERS, build_forest, dfs_order  # noqa: E402

OUT = Path(__file__).resolve().parent.parent / "output" / "树状排序示例.xlsx"
TMP_DIR = Path(__file__).resolve().parent.parent / "output" / "_sort_demo_tmp"

# 故意打乱行序：祖先不一定在前，子树被拆散；#100/#101 也夹在中间/偏后
# 逻辑路径 category_pids 仍然正确 —— 排序只靠路径，不靠当前行序
# 同父兄弟的相对顺序 = 本表中该子节点首次出现顺序（故下面让 稻谷→小麦→玉米 先被扫到）
DISORDERED_ROWS = [
    (13, "硬质小麦", "2,3,4,12", "[-1],[2],[3],[4],[12]", "农林牧渔,农业产品,谷物,小麦", "[]"),
    (7, "晚籼稻", "2,3,4,5", "[-1],[2],[3],[4],[5]", "农林牧渔,农业产品,谷物,稻谷", "[]"),
    (14, "软质小麦", "2,3,4,12", "[-1],[2],[3],[4],[12]", "农林牧渔,农业产品,谷物,小麦", "[]"),
    (6, "早籼稻", "2,3,4,5", "[-1],[2],[3],[4],[5]", "农林牧渔,农业产品,谷物,稻谷", "[]"),
    (101, "糯玉米", "2,3,4,16", "[-1],[2],[3],[4],[16]", "农林牧渔,农业产品,谷物,玉米", "[]"),
    (3, "农业产品", "2", "[-1],[2]", "农林牧渔", "[]"),
    (5, "稻谷", "2,3,4", "[-1],[2],[3],[4]", "农林牧渔,农业产品,谷物", "[]"),
    (100, "新型小麦", "2,3,4,12", "[-1],[2],[3],[4],[12]", "农林牧渔,农业产品,谷物,小麦", "[]"),
    (12, "小麦", "2,3,4", "[-1],[2],[3],[4]", "农林牧渔,农业产品,谷物", "[]"),
    (16, "玉米", "2,3,4", "[-1],[2],[3],[4]", "农林牧渔,农业产品,谷物", '["苞米"]'),
    (2, "农林牧渔", None, "[-1],", None, "[]"),
    (4, "谷物", "2,3", "[-1],[2],[3]", "农林牧渔,农业产品", "[]"),
]

# 排序后期望的先序（用于「说明」列注释；实际行序以 DFS 为准）
DEPTH_NOTE = {
    "2": "根",
    "3": "└ 农业产品",
    "4": "   └ 谷物",
    "5": "      └ 稻谷",
    "6": "         └ 叶",
    "7": "         └ 叶",
    "12": "      └ 小麦",
    "13": "         └ 叶",
    "14": "         └ 叶",
    "100": "         └ 叶（原在表尾，排回小麦下）",
    "16": "      └ 玉米",
    "101": "         └ 叶（原在表尾，排回玉米下）",
}


def _style_header(ws) -> None:
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F5D50")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")


def _apply_borders(ws) -> None:
    thin = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = thin


def _set_widths(ws, widths: list[int]) -> None:
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = width


def write_raw_disorder(path: Path) -> None:
    """写出仅含正式 6 列的乱序表，供排序逻辑读取。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(HEADERS)
    for row in DISORDERED_ROWS:
        ws.append(list(row))
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def load_category_rows(path: Path) -> list[CategoryRow]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    header = [str(h).strip() if h is not None else "" for h in next(it)]
    col = {name: i for i, name in enumerate(header)}
    rows: list[CategoryRow] = []
    for order, row in enumerate(it):
        cid = str(row[col["category_id"]]).strip()
        cname = str(row[col["category_name"]]).strip()
        rows.append(
            CategoryRow(
                category_id=cid,
                category_name=cname,
                category_group_id=row[col.get("category_group_id", -1)] if "category_group_id" in col else None,
                category_pids=row[col.get("category_pids", -1)] if "category_pids" in col else None,
                category_group_name=row[col.get("category_group_name", -1)] if "category_group_name" in col else None,
                syn_list=row[col.get("syn_list", -1)] if "syn_list" in col else "[]",
                source_order=order,
            )
        )
    wb.close()
    return rows


def build_presentation(before: list[tuple], after: list[CategoryRow], out: Path) -> None:
    wb = Workbook()
    guide = wb.active
    guide.title = "阅读说明"
    guide["A1"] = "树状排序示例：乱序 → 排序 → 有序"
    guide["A1"].font = Font(bold=True, size=14)
    guide["A3"] = "本文件专门演示 sort_taxonomy_excel.py 的效果（不是树状插入）。"
    guide["A4"] = "Sheet「排序前_乱序」：行序被打乱，子树拆散，#100/#101 堆在表尾；category_pids 仍正确。"
    guide["A5"] = "Sheet「排序后_有序」：按 pids 建树后做先序 DFS，父在上、整棵子树紧跟；末尾追加节点也回到父下。"
    guide["A7"] = "对比要点："
    guide["A8"] = "· 排序前：叶子跑到父亲前面；根/谷物偏后；#100/#101 不在父节点子树旁。"
    guide["A9"] = "· 排序后：农林牧渔→农业产品→谷物→稻谷子树→小麦子树(含新型小麦)→玉米子树(含糯玉米)"
    guide["A11"] = "对应命令："
    guide["A12"] = "python scripts/sort_taxonomy_excel.py --input <乱序表.xlsx> --output <有序表.xlsx>"
    guide.column_dimensions["A"].width = 100

    present_headers = HEADERS + ["说明"]
    yellow = PatternFill("solid", fgColor="FFF2CC")
    green = PatternFill("solid", fgColor="E2EFDA")

    # —— 排序前 ——
    ws1 = wb.create_sheet("排序前_乱序")
    ws1.append(present_headers)
    _style_header(ws1)
    disorder_notes = {
        13: "叶子跑到父亲前面",
        7: "叶子跑到父亲前面",
        14: "叶子位置散乱",
        6: "叶子散落",
        101: "← 应属玉米，却远离父节点",
        3: "中间层位置乱",
        5: "与子女不在一起",
        100: "← 应属小麦，却远离父节点",
        12: "小麦与子女分离",
        16: "玉米与糯玉米分离",
        2: "根不在最上",
        4: "谷物偏到最后",
    }
    for row in before:
        cid = row[0]
        note = disorder_notes.get(cid, "行序错乱")
        ws1.append(list(row) + [note])
        if cid in (100, 101, 13, 7, 2, 4):
            for cell in ws1[ws1.max_row]:
                cell.fill = yellow
    _apply_borders(ws1)
    _set_widths(ws1, [12, 14, 16, 28, 40, 12, 34])

    # —— 排序后 ——
    ws2 = wb.create_sheet("排序后_有序")
    ws2.append(present_headers)
    _style_header(ws2)
    for r in after:
        note = DEPTH_NOTE.get(r.category_id, "树序")
        ws2.append(list(r.as_tuple()) + [note])
        if r.category_id in ("100", "101"):
            for cell in ws2[ws2.max_row]:
                cell.fill = green
    _apply_borders(ws2)
    _set_widths(ws2, [12, 14, 16, 28, 40, 12, 34])

    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)


def main() -> None:
    TMP_DIR.mkdir(parents=True, exist_ok=True)
    raw = TMP_DIR / "disorder_input.xlsx"
    write_raw_disorder(raw)

    rows = load_category_rows(raw)
    ordered = dfs_order(build_forest(rows))

    build_presentation(DISORDERED_ROWS, ordered, OUT)
    print(f"written: {OUT}")
    print("排序前行序:", [r[0] for r in DISORDERED_ROWS])
    print("排序后行序:", [r.category_id for r in ordered])


if __name__ == "__main__":
    main()
